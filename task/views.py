import time
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse
# Create your views here.
from django.views.generic.base import TemplateView
import concurrent.futures
from task.models import Applicants


class ExcelPageView(TemplateView):
    template_name = "excel_page.html"


def ExportThread(app_data, columns, sheet, row_num):
    for app in app_data:
        row_num += 1
        row = [
            app['id'],
            app['job__title'],
            app['job__category__title'],
            app['user__username'],
            app['documents__category'],
            app['documents__file__type']
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)


def ExportExcelfile(request):

    queryset = Applicants.objects.values(
        'id', 'job__title', 'job__category__title', 'user__username', 'documents__category', 'documents__file__type')

    wb = Workbook()
    wb.title = "Database data"
    sheet = wb.active

    columns = ['id', 'Job', 'Job Category',
               'Applicant', 'Document', 'file']

    for col_num, column_title in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # no of threads
    n_workers = 2

    chunksize = int((len(queryset) / n_workers))

    # starttime
    start = time.time()
    with concurrent.futures.ThreadPoolExecutor(n_workers) as executor:
        # split operations into chunks
        for i in range(1, len(queryset), chunksize):
            row_num = i
            # select a chunk
            app_data = queryset[i:(i + chunksize)]
            # submit the batch
            k = executor.map(
                ExportThread(app_data, columns, sheet, row_num))

    # endtime
    end = time.time()
    # totaltime
    print(f'Total time to run: {end - start:2f}s')
    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=exportedfile.xlsx'
    wb.save(response)
    return response
